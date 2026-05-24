"""
generate_templates.py — Script to build the three starter .xlsx template files.

Run once during docker image build (or locally to regenerate):
    python generate_templates.py

Outputs:
    summary.xlsx       -- Cover + Data + Charts sheets
    daily-report.xlsx  -- Single-sheet report with KPI section
    comparison.xlsx    -- Side-by-side comparison with conditional formatting
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import PatternFill as PF


HERE = Path(__file__).parent


def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill(patternType="solid", fgColor=hex_color)


def _bold_font(size: int = 11, color: str = "FFFFFF", name: str = "Arial") -> Font:
    return Font(name=name, size=size, bold=True, color=color)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _thin_border() -> Border:
    thin = Side(style="thin", color="D1D5DB")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ---------------------------------------------------------------------------
# summary.xlsx
# ---------------------------------------------------------------------------

def make_summary() -> None:
    """Three-sheet template: Cover / Data / Charts."""
    wb = openpyxl.Workbook()

    # ----- Cover sheet -----
    cover = wb.active
    cover.title = "封面"

    cover["A1"] = "{{title}}"
    cover["A1"].font = Font(name="Arial", size=20, bold=True, color="1F2937")
    cover["A1"].alignment = _center()

    cover["A2"] = "报告周期：{{period}}"
    cover["A2"].font = Font(name="Arial", size=12, color="6B7280")
    cover["A2"].alignment = _center()

    cover["A3"] = ""

    cover["A4"] = "报告日期：{{generated_date}}"
    cover["A4"].font = Font(name="Arial", size=11, color="374151")
    cover["A4"].alignment = Alignment(horizontal="left", vertical="center")

    cover["A5"] = "摘要：{{summary_text}}"
    cover["A5"].font = Font(name="Arial", size=11, color="374151")
    cover["A5"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cover["A5"].fill = _header_fill("EFF6FF")

    cover.row_dimensions[1].height = 40
    cover.row_dimensions[2].height = 25
    cover.row_dimensions[5].height = 60
    cover.column_dimensions["A"].width = 60

    # ----- Data sheet -----
    data_ws = wb.create_sheet(title="数据")

    headers = ["项目", "1月", "2月", "3月", "合计"]
    placeholder_rows = [
        ["数据项 A", 0, 0, 0, 0],
        ["数据项 B", 0, 0, 0, 0],
        ["数据项 C", 0, 0, 0, 0],
    ]

    for col_idx, h in enumerate(headers, start=1):
        cell = data_ws.cell(row=1, column=col_idx, value=h)
        cell.fill = _header_fill("2563EB")
        cell.font = _bold_font(10)
        cell.alignment = _center()

    for row_idx, row in enumerate(placeholder_rows, start=2):
        for col_idx, val in enumerate(row, start=1):
            cell = data_ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _thin_border()
            if row_idx % 2 == 0:
                cell.fill = _header_fill("EFF6FF")

    data_ws.freeze_panes = "A2"
    for i, w in enumerate([20, 10, 10, 10, 12], start=1):
        from openpyxl.utils import get_column_letter
        data_ws.column_dimensions[get_column_letter(i)].width = w

    # ----- Charts sheet -----
    charts_ws = wb.create_sheet(title="图表")
    charts_ws["A1"] = "（图表将由 xlsx-author skill 在此处插入）"
    charts_ws["A1"].font = Font(name="Arial", size=10, color="9CA3AF", italic=True)

    wb.save(str(HERE / "summary.xlsx"))
    print("Created: summary.xlsx")


# ---------------------------------------------------------------------------
# daily-report.xlsx
# ---------------------------------------------------------------------------

def make_daily_report() -> None:
    """Single-sheet daily/weekly report."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "日报"

    # Row 1-3: Header area
    ws.merge_cells("A1:H1")
    ws["A1"] = "{{report_date}} {{report_type}}报告 — {{department}}"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = _header_fill("1F2937")
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 45

    ws.merge_cells("A2:H2")
    ws["A2"] = "部门：{{department}}　|　制表人：{{author}}　|　生成时间：{{generated_date}}"
    ws["A2"].font = Font(name="Arial", size=10, color="6B7280")
    ws["A2"].fill = _header_fill("F3F4F6")
    ws["A2"].alignment = _center()
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 10  # spacer

    # Row 4: KPI section label
    ws.merge_cells("A4:H4")
    ws["A4"] = "— 核心 KPI —"
    ws["A4"].font = Font(name="Arial", size=11, bold=True, color="374151")
    ws["A4"].fill = _header_fill("E5E7EB")
    ws["A4"].alignment = _center()
    ws.row_dimensions[4].height = 25

    # Rows 5-8: KPI card placeholders (4 KPIs, 2 per row)
    kpi_labels = ["KPI 指标 A", "KPI 指标 B", "KPI 指标 C", "KPI 指标 D"]
    kpi_positions = [("A5", "B7"), ("D5", "E7"), ("A8", "B10"), ("D8", "E10")]
    for label, (start, end) in zip(kpi_labels, kpi_positions):
        ws.merge_cells(f"{start}:{end}")
        cell = ws[start]
        cell.value = f"【{label}】\n0"
        cell.font = Font(name="Arial", size=14, bold=True, color="1D4ED8")
        cell.fill = _header_fill("DBEAFE")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[11].height = 10  # spacer

    # Row 12 onwards: data table placeholder
    ws.merge_cells("A12:H12")
    ws["A12"] = "— 明细数据 —"
    ws["A12"].font = Font(name="Arial", size=11, bold=True, color="374151")
    ws["A12"].fill = _header_fill("E5E7EB")
    ws["A12"].alignment = _center()

    table_headers = ["序号", "项目名称", "本日数量", "本周累计", "本月累计", "目标值", "完成率", "备注"]
    for col_idx, h in enumerate(table_headers, start=1):
        cell = ws.cell(row=13, column=col_idx, value=h)
        cell.fill = _header_fill("374151")
        cell.font = _bold_font(10)
        cell.alignment = _center()

    ws.freeze_panes = "A14"

    # Column widths
    widths = [6, 20, 12, 12, 12, 12, 10, 16]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(str(HERE / "daily-report.xlsx"))
    print("Created: daily-report.xlsx")


# ---------------------------------------------------------------------------
# comparison.xlsx
# ---------------------------------------------------------------------------

def make_comparison() -> None:
    """Side-by-side comparison template with conditional formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对比分析"

    # Title row
    ws.merge_cells("A1:K1")
    ws["A1"] = "{{metric_name}} — 对比分析"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = _header_fill("1F2937")
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 35

    # Header row 2
    # Group A: columns A-D (base group)
    ws.merge_cells("A2:D2")
    ws["A2"] = "{{base_label}}（基准）"
    ws["A2"].font = _bold_font(10, "FFFFFF")
    ws["A2"].fill = _header_fill("2563EB")
    ws["A2"].alignment = _center()

    # Spacer column E
    ws["E2"] = ""
    ws["E2"].fill = _header_fill("F3F4F6")

    # Group B: columns F-I (compare group)
    ws.merge_cells("F2:I2")
    ws["F2"] = "{{compare_label}}（对比）"
    ws["F2"].font = _bold_font(10, "FFFFFF")
    ws["F2"].fill = _header_fill("16A34A")
    ws["F2"].alignment = _center()

    # Diff column
    ws["J2"] = "差值"
    ws["J2"].font = _bold_font(10, "FFFFFF")
    ws["J2"].fill = _header_fill("7C3AED")
    ws["J2"].alignment = _center()

    ws["K2"] = "变化率"
    ws["K2"].font = _bold_font(10, "FFFFFF")
    ws["K2"].fill = _header_fill("7C3AED")
    ws["K2"].alignment = _center()

    # Sub-headers row 3
    base_subs = ["指标", "数值", "单位", "备注"]
    cmp_subs = ["指标", "数值", "单位", "备注"]
    cols_A = ["A", "B", "C", "D"]
    cols_F = ["F", "G", "H", "I"]

    for col_letter, label in zip(cols_A, base_subs):
        cell = ws[f"{col_letter}3"]
        cell.value = label
        cell.font = Font(name="Arial", size=9, bold=True, color="374151")
        cell.fill = _header_fill("DBEAFE")
        cell.alignment = _center()

    ws["E3"] = ""

    for col_letter, label in zip(cols_F, cmp_subs):
        cell = ws[f"{col_letter}3"]
        cell.value = label
        cell.font = Font(name="Arial", size=9, bold=True, color="374151")
        cell.fill = _header_fill("DCFCE7")
        cell.alignment = _center()

    ws["J3"] = "差值(对比-基准)"
    ws["J3"].font = Font(name="Arial", size=9, bold=True, color="374151")
    ws["J3"].fill = _header_fill("EDE9FE")
    ws["J3"].alignment = _center()

    ws["K3"] = "(对比-基准)/基准"
    ws["K3"].font = Font(name="Arial", size=9, bold=True, color="374151")
    ws["K3"].fill = _header_fill("EDE9FE")
    ws["K3"].alignment = _center()

    # Placeholder data rows (4-8)
    placeholder_items = ["项目 1", "项目 2", "项目 3", "项目 4", "合计"]
    for row_offset, item in enumerate(placeholder_items):
        row_idx = 4 + row_offset
        bg = "EFF6FF" if row_offset % 2 == 0 else "FFFFFF"
        ws.cell(row=row_idx, column=1, value=item).fill = _header_fill(bg)
        ws.cell(row=row_idx, column=2, value=0).fill = _header_fill(bg)
        ws.cell(row=row_idx, column=3, value="").fill = _header_fill(bg)
        ws.cell(row=row_idx, column=4, value="").fill = _header_fill(bg)

        ws.cell(row=row_idx, column=5, value="")  # spacer

        bg2 = "F0FDF4" if row_offset % 2 == 0 else "FFFFFF"
        ws.cell(row=row_idx, column=6, value=item).fill = _header_fill(bg2)
        ws.cell(row=row_idx, column=7, value=0).fill = _header_fill(bg2)
        ws.cell(row=row_idx, column=8, value="").fill = _header_fill(bg2)
        ws.cell(row=row_idx, column=9, value="").fill = _header_fill(bg2)

        # Diff formula placeholder (=G{n}-B{n})
        diff_cell = ws.cell(row=row_idx, column=10, value=f"=G{row_idx}-B{row_idx}")
        rate_cell = ws.cell(row=row_idx, column=11, value=f"=IF(B{row_idx}=0,\"\",J{row_idx}/B{row_idx})")
        rate_cell.number_format = "0.0%"

    # Conditional formatting on diff column (J4:J8): green positive, red negative.
    from openpyxl.formatting.rule import CellIsRule
    green_fill = PF(patternType="solid", fgColor="D1FAE5")
    red_fill = PF(patternType="solid", fgColor="FEE2E2")

    ws.conditional_formatting.add(
        "J4:J8",
        CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill),
    )
    ws.conditional_formatting.add(
        "J4:J8",
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill),
    )

    ws.freeze_panes = "A4"

    # Column widths
    col_widths = {
        "A": 14, "B": 10, "C": 8, "D": 12,
        "E": 3,
        "F": 14, "G": 10, "H": 8, "I": 12,
        "J": 15, "K": 12,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    wb.save(str(HERE / "comparison.xlsx"))
    print("Created: comparison.xlsx")


if __name__ == "__main__":
    make_summary()
    make_daily_report()
    make_comparison()
    print("All templates generated successfully.")

"""
style.py — Cell styling helpers for xlsx-author skill.

All fill_color / font_color arguments accept hex strings WITHOUT the leading '#'
(e.g. "2563EB", not "#2563EB"). This matches openpyxl's PatternFill convention.

Functions:
    apply_header_style   -- fill + bold font for the header row
    apply_alt_row_fill   -- alternating row background colour
    auto_column_width    -- auto-size columns to content
    apply_number_format  -- set number format for a whole column
    sanitize_sheet_name  -- clean sheet names (31-char limit, illegal chars removed)
    THEME_PRESETS        -- dict of four built-in colour themes
"""

from __future__ import annotations

import re
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Theme presets
# ---------------------------------------------------------------------------

#: Four built-in colour themes.  ``header_fill`` and ``alt_row_fill`` are hex
#: strings without '#'; ``font_color`` is the header text colour.
THEME_PRESETS: dict[str, dict[str, str]] = {
    "default": {
        "header_fill": "1F2937",   # dark grey
        "alt_row_fill": "F9FAFB",  # very light grey
        "font_color": "FFFFFF",
    },
    "blue": {
        "header_fill": "2563EB",   # brand blue
        "alt_row_fill": "EFF6FF",  # ice blue
        "font_color": "FFFFFF",
    },
    "green": {
        "header_fill": "16A34A",   # success green
        "alt_row_fill": "F0FDF4",  # mint
        "font_color": "FFFFFF",
    },
    "monochrome": {
        "header_fill": "374151",   # mid grey
        "alt_row_fill": "F3F4F6",  # near-white
        "font_color": "FFFFFF",
    },
}


# ---------------------------------------------------------------------------
# Public helpers
# ---------------------------------------------------------------------------

def apply_header_style(
    ws: Worksheet,
    row: int = 1,
    fill_color: str = "2563EB",
    font_color: str = "FFFFFF",
    font_name: str = "微软雅黑",
    font_size: int = 10,
    bold: bool = True,
) -> None:
    """Apply header styling (fill + bold font + centre alignment) to *row*.

    Args:
        ws:          Target worksheet.
        row:         1-based row index to style (default: 1).
        fill_color:  Hex fill colour for the header cells, WITHOUT '#' prefix.
                     Example: "2563EB" for brand blue.
        font_color:  Hex font colour, WITHOUT '#' prefix. Default white.
        font_name:   Font family name.  "微软雅黑" renders well on Windows;
                     macOS/Linux fall back to the system CJK font gracefully.
        font_size:   Point size (default 10).
        bold:        Whether the header text is bold (default True).

    Note:
        openpyxl does NOT embed font files into the xlsx.  The font_name is
        stored as a string; the actual rendering depends on the viewer's
        installed fonts.
    """
    fill = PatternFill(patternType="solid", fgColor=fill_color)
    font = Font(
        name=font_name,
        size=font_size,
        bold=bold,
        color=font_color,
    )
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def apply_alt_row_fill(
    ws: Worksheet,
    start_row: int = 2,
    end_row: Optional[int] = None,
    even_fill: str = "F8FAFC",
) -> None:
    """Apply a light background fill to even-numbered data rows.

    Odd rows (1-based after start_row) remain transparent; even rows receive
    ``even_fill``.  This creates the classic "striped table" effect.

    Args:
        ws:         Target worksheet.
        start_row:  First data row (1-based).  Row 1 is typically the header.
        end_row:    Last data row (inclusive).  ``None`` means auto-detect from
                    ``ws.max_row``.
        even_fill:  Hex fill colour for even rows, WITHOUT '#' prefix.
    """
    if end_row is None:
        end_row = ws.max_row

    fill = PatternFill(patternType="solid", fgColor=even_fill)

    for row_idx in range(start_row, end_row + 1):
        # Apply fill to even rows (2nd, 4th, 6th… relative to the full sheet).
        if (row_idx - start_row) % 2 == 1:
            for cell in ws[row_idx]:
                cell.fill = fill


def auto_column_width(
    ws: Worksheet,
    min_width: int = 8,
    max_width: int = 50,
) -> None:
    """Automatically size each column to fit its longest cell value.

    The width is computed as ``max_content_length * 1.2`` clamped to
    ``[min_width, max_width]``.  For CJK characters each character counts as
    2 units (approximation; actual rendering depends on the font).

    Args:
        ws:         Target worksheet.
        min_width:  Minimum column width in character units.
        max_width:  Maximum column width in character units.

    Note:
        openpyxl column widths are in "character units" measured in the default
        font.  This approximation works well for typical business data.
    """
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is None:
                continue
            text = str(cell.value)
            # CJK characters: count as 2 units each (rough approximation).
            length = sum(2 if ord(c) > 0x2E7F else 1 for c in text)
            max_len = max(max_len, length)
        adjusted = min(max(int(max_len * 1.2), min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def apply_number_format(
    ws: Worksheet,
    col_index: int,
    fmt: str = "#,##0.00",
    start_row: int = 2,
) -> None:
    """Apply a number format string to all data cells in a column.

    Args:
        ws:         Target worksheet.
        col_index:  1-based column index.
        fmt:        Excel number format string (e.g. "#,##0.00", "0%",
                    "YYYY-MM-DD").
        start_row:  First row to format (skip header rows).
    """
    for row_idx in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_index)
        if cell.value is not None:
            cell.number_format = fmt


def sanitize_sheet_name(name: str) -> str:
    """Sanitise an Excel sheet name.

    Removes characters forbidden by Excel (``[ ] : * ? / \\``) and truncates
    to 31 characters (Excel's hard limit, measured in Unicode code points, not
    bytes).

    Args:
        name:  Raw sheet name (may contain illegal chars or be too long).

    Returns:
        A cleaned sheet name safe for openpyxl.
    """
    # Remove illegal characters
    cleaned = re.sub(r"[\[\]:*?/\\]", "", name)
    # Collapse runs of whitespace
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    # Truncate to 31 characters
    return cleaned[:31] if cleaned else "Sheet"

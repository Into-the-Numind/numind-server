"""
csv_reader.py — CSV → openpyxl worksheet helper for xlsx-author skill.

Reads a CSV file (UTF-8 or UTF-8-BOM encoding) using pandas, cleans NaN
values so openpyxl does not crash, and writes the data as a new worksheet
in an existing Workbook.

Usage::

    import openpyxl
    from helpers.csv_reader import csv_to_sheet

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    csv_to_sheet(wb, "/workspace/input/sales.csv", sheet_name="销售数据")
    wb.save("/output/report.xlsx")
"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.workbook.workbook import Workbook

from .style import sanitize_sheet_name


def csv_to_sheet(
    wb: Workbook,
    csv_path: str,
    sheet_name: Optional[str] = None,
    infer_types: bool = True,
) -> str:
    """Read a CSV file and append it as a new worksheet in *wb*.

    NaN / infinity values are converted to Python ``None`` before writing so
    that openpyxl does not raise ``ValueError`` on non-finite floats.

    Datetime columns (``dtype == datetime64``) are converted to ISO-format
    strings (``YYYY-MM-DD``) to avoid openpyxl timezone handling issues.

    Args:
        wb:           Target openpyxl ``Workbook``.
        csv_path:     Absolute path to the CSV file.
        sheet_name:   Worksheet name.  Defaults to the CSV file stem (first 31
                      chars, sanitised for Excel).
        infer_types:  When True, pandas infers column dtypes.  Set False to
                      read everything as strings (safe fallback for mixed columns).

    Returns:
        The actual sheet name used (may differ from ``sheet_name`` after
        sanitisation or truncation).

    Raises:
        FileNotFoundError: When *csv_path* does not exist.
        pd.errors.ParserError: When the CSV is malformed.
    """
    path = Path(csv_path)
    if not path.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_path}")

    # UTF-8-BOM (Excel default on Windows) is handled by encoding="utf-8-sig".
    df = pd.read_csv(str(path), encoding="utf-8-sig", dtype=None if infer_types else str)

    # Sanitise sheet name.
    if sheet_name is None:
        sheet_name = sanitize_sheet_name(path.stem)
    else:
        sheet_name = sanitize_sheet_name(sheet_name)

    ws = wb.create_sheet(title=sheet_name)

    # Write header row.
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=str(col_name))

    # Convert datetime columns to strings before writing.
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.strftime("%Y-%m-%d")

    # Write data rows, replacing NaN/Inf with None.
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row_data, start=1):
            # Convert float NaN / Inf to None (openpyxl cannot handle them).
            if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
                val = None
            # pandas NA
            if pd.isna(val) if not isinstance(val, (list, dict)) else False:
                val = None
            ws.cell(row=row_idx, column=col_idx, value=val)

    return sheet_name

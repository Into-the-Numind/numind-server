"""
example_with_chart.py — Advanced xlsx-author skill usage example.

Demonstrates:
- Multiple sheets (data + chart + summary)
- Embedded bar chart (openpyxl native) on a dedicated chart sheet
- Embedded line chart for trend data
- matplotlib chart embedded as PNG (WPS-compatible fallback)
- CSV input reading from /workspace/input/
- "green" colour theme with number formatting
- Template variable substitution using the summary template

Run inside the sandbox (with test CSV):
    cd /skills/xlsx-author

    # Create a test CSV first (optional — skipped if file not found)
    echo "产品,Q1销量,Q2销量,Q3销量" > /workspace/input/products.csv
    echo "产品A,1200,1500,1800" >> /workspace/input/products.csv
    echo "产品B,800,950,1100" >> /workspace/input/products.csv

    python examples/example_with_chart.py

Output: /output/example_with_chart.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path

# Make helpers importable when running directly.
_SKILL_DIR = Path(__file__).parent.parent
if str(_SKILL_DIR) not in sys.path:
    sys.path.insert(0, str(_SKILL_DIR))

import openpyxl
from helpers import (
    apply_header_style,
    apply_alt_row_fill,
    auto_column_width,
    apply_number_format,
    THEME_PRESETS,
)
from helpers.chart import embed_bar_chart, embed_line_chart, chart_from_matplotlib
from helpers.csv_reader import csv_to_sheet

OUTPUT_DIR = Path("/output")
INPUT_DIR = Path("/workspace/input")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def main() -> None:
    theme = THEME_PRESETS["green"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---- Sheet 1: Monthly revenue data ----
    ws_data = wb.create_sheet(title="月度数据")

    revenue_data = [
        ["月份", "营收（万元）", "成本（万元）", "毛利润（万元）", "毛利率"],
        ["2026-01",  52.3,  31.5,  20.8, 0.398],
        ["2026-02",  48.7,  29.8,  18.9, 0.388],
        ["2026-03",  61.5,  36.2,  25.3, 0.412],
        ["2026-04",  71.2,  41.0,  30.2, 0.424],
        ["2026-05",  68.9,  39.5,  29.4, 0.427],
        ["2026-06",  79.4,  44.1,  35.3, 0.445],
    ]

    for row in revenue_data:
        ws_data.append(row)

    apply_header_style(
        ws_data, row=1,
        fill_color=theme["header_fill"],
        font_color=theme["font_color"],
    )
    apply_alt_row_fill(ws_data, start_row=2, even_fill=theme["alt_row_fill"])
    apply_number_format(ws_data, col_index=2, fmt="#,##0.0")
    apply_number_format(ws_data, col_index=3, fmt="#,##0.0")
    apply_number_format(ws_data, col_index=4, fmt="#,##0.0")
    apply_number_format(ws_data, col_index=5, fmt="0.0%")
    auto_column_width(ws_data)
    ws_data.freeze_panes = "A2"

    # ---- Sheet 2: openpyxl native charts ----
    ws_charts = wb.create_sheet(title="图表（openpyxl）")
    ws_charts["A1"] = "月度收支对比图（openpyxl 原生图表，Excel 内可交互）"
    ws_charts["A1"].font = openpyxl.styles.Font(italic=True, color="6B7280", size=9)

    # Bar chart: revenue vs cost
    embed_bar_chart(
        wb=wb,
        data_sheet_name="月度数据",
        data_range="A1:D7",
        title="月度营收 / 成本 / 毛利润对比",
        anchor_cell="A3",
        target_sheet_name="图表（openpyxl）",
        series_labels=["营收（万元）", "成本（万元）", "毛利润（万元）"],
        width_cm=18,
        height_cm=12,
    )

    # Line chart: gross margin trend (columns A + E)
    embed_line_chart(
        wb=wb,
        data_sheet_name="月度数据",
        data_range="A1:A7",    # categories
        title="毛利率趋势",
        anchor_cell="P3",
        target_sheet_name="图表（openpyxl）",
        smooth=True,
        width_cm=14,
        height_cm=10,
    )

    # ---- Sheet 3: matplotlib chart (WPS-compatible PNG) ----
    ws_mpl = wb.create_sheet(title="图表（matplotlib）")
    ws_mpl["A1"] = "月度毛利率趋势图（matplotlib 渲染，全平台兼容）"
    ws_mpl["A1"].font = openpyxl.styles.Font(italic=True, color="6B7280", size=9)

    try:
        import matplotlib
        matplotlib.rcParams["font.family"] = ["WenQuanYi Zen Hei", "DejaVu Sans"]
        import matplotlib.pyplot as plt

        months = ["1月", "2月", "3月", "4月", "5月", "6月"]
        margins = [0.398, 0.388, 0.412, 0.424, 0.427, 0.445]

        fig, ax = plt.subplots(figsize=(9, 5))
        ax.plot(months, [m * 100 for m in margins], marker="o",
                color="#16A34A", linewidth=2, markersize=7)
        ax.fill_between(range(len(months)),
                        [m * 100 for m in margins],
                        alpha=0.15, color="#16A34A")
        ax.set_title("月度毛利率趋势（2026 H1）", fontsize=13)
        ax.set_ylabel("毛利率（%）", fontsize=10)
        ax.set_ylim(30, 55)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.1f}%"))
        ax.grid(axis="y", linestyle="--", alpha=0.4)
        plt.tight_layout()

        chart_from_matplotlib(fig, ws_mpl, anchor_cell="A3", dpi=150)
        plt.close(fig)
        print("matplotlib chart embedded successfully.")
    except ImportError as e:
        ws_mpl["A2"] = f"（matplotlib not available: {e}）"

    # ---- Sheet 4: CSV input (optional) ----
    csv_path = INPUT_DIR / "products.csv"
    if csv_path.exists():
        csv_to_sheet(wb, str(csv_path), sheet_name="产品数据")
        ws_csv = wb["产品数据"]
        apply_header_style(ws_csv, row=1,
            fill_color=theme["header_fill"],
            font_color=theme["font_color"])
        apply_alt_row_fill(ws_csv, start_row=2, even_fill=theme["alt_row_fill"])
        auto_column_width(ws_csv)
        print(f"CSV sheet added from {csv_path}")
    else:
        print(f"Skipping CSV sheet: {csv_path} not found.")

    # ---- Save ----
    output_path = OUTPUT_DIR / "example_with_chart.xlsx"
    wb.save(str(output_path))
    size_kb = output_path.stat().st_size / 1024
    print(f"Saved: {output_path} ({size_kb:.1f} KB)")
    print(f"Sheets: {[s.title for s in wb.worksheets]}")


if __name__ == "__main__":
    main()

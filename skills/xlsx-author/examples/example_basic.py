"""
example_basic.py — Basic xlsx-author skill usage example.

Demonstrates:
- Creating a single-sheet workbook
- Applying the "blue" theme (header fill + alternating row fill)
- Auto-sizing columns
- Freezing the top row
- Saving to /output/

Run inside the sandbox:
    cd /skills/xlsx-author
    python examples/example_basic.py

Output: /output/example_basic.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path

# Make helpers importable when running directly from the examples/ directory.
_SKILL_DIR = Path(__file__).parent.parent
if str(_SKILL_DIR) not in sys.path:
    sys.path.insert(0, str(_SKILL_DIR))

import openpyxl
from helpers import apply_header_style, apply_alt_row_fill, auto_column_width, THEME_PRESETS

OUTPUT_DIR = Path("/output")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def main() -> None:
    # ---- 1. Create workbook ----
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default "Sheet"

    # ---- 2. Write data ----
    ws = wb.create_sheet(title="销售月报")

    data = [
        ["月份",   "区域",     "销售额（元）", "订单数", "均单价（元）", "完成率"],
        ["2026-01", "华东区",   1_200_000,     340,      3529.41,      0.95],
        ["2026-01", "华南区",     980_000,     280,      3500.00,      0.88],
        ["2026-01", "华北区",   1_050_000,     310,      3387.10,      0.92],
        ["2026-02", "华东区",   1_150_000,     325,      3538.46,      0.91],
        ["2026-02", "华南区",     920_000,     265,      3471.70,      0.83],
        ["2026-02", "华北区",     995_000,     295,      3372.88,      0.87],
        ["2026-03", "华东区",   1_380_000,     390,      3538.46,      1.09],
        ["2026-03", "华南区",   1_100_000,     310,      3548.39,      0.99],
        ["2026-03", "华北区",   1_240_000,     355,      3492.96,      1.09],
    ]

    for row in data:
        ws.append(row)

    # ---- 3. Apply blue theme styling ----
    theme = THEME_PRESETS["blue"]
    apply_header_style(
        ws,
        row=1,
        fill_color=theme["header_fill"],
        font_color=theme["font_color"],
        font_name="微软雅黑",
        font_size=10,
        bold=True,
    )
    apply_alt_row_fill(ws, start_row=2, even_fill=theme["alt_row_fill"])

    # ---- 4. Number formatting ----
    from helpers import apply_number_format
    apply_number_format(ws, col_index=3, fmt="#,##0")        # 销售额
    apply_number_format(ws, col_index=5, fmt="#,##0.00")     # 均单价
    apply_number_format(ws, col_index=6, fmt="0%")           # 完成率

    # ---- 5. Auto column width + freeze ----
    auto_column_width(ws)
    ws.freeze_panes = "A2"

    # ---- 6. Save ----
    output_path = OUTPUT_DIR / "example_basic.xlsx"
    wb.save(str(output_path))
    print(f"Saved: {output_path} ({output_path.stat().st_size:,} bytes)")
    print(f"Sheets: {[s.title for s in wb.worksheets]}")
    print(f"Rows:   {ws.max_row - 1} data rows (excluding header)")


if __name__ == "__main__":
    main()

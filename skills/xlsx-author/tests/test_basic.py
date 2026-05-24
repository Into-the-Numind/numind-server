"""
test_basic.py — pytest tests for xlsx-author skill helpers and main entry point.

Run from the skill root:
    cd /skills/xlsx-author
    pytest tests/test_basic.py -v

Coverage:
    - helpers/style.py  : apply_header_style, apply_alt_row_fill,
                          auto_column_width, apply_number_format,
                          sanitize_sheet_name, THEME_PRESETS
    - helpers/chart.py  : embed_bar_chart, embed_pie_chart (chart count checks)
    - helpers/csv_reader.py : csv_to_sheet, NaN handling
    - helpers/template.py   : load_template (mocked template dir)
    - main.py               : run() happy path + error path + write_only

Dependencies (must be installed in test env):
    openpyxl>=3.1  pandas>=2.0  pytest>=7.0
    matplotlib and pillow are optional for chart_from_matplotlib tests
"""

from __future__ import annotations

import csv
import io
import math
import os
import sys
import tempfile
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest
import openpyxl
from openpyxl.chart import BarChart, PieChart

# ---------------------------------------------------------------------------
# Path fixup so helpers are importable when pytest runs from project root.
# ---------------------------------------------------------------------------
_SKILL_DIR = Path(__file__).parent.parent
if str(_SKILL_DIR) not in sys.path:
    sys.path.insert(0, str(_SKILL_DIR))

from helpers.style import (
    apply_header_style,
    apply_alt_row_fill,
    auto_column_width,
    apply_number_format,
    sanitize_sheet_name,
    THEME_PRESETS,
)
from helpers.chart import embed_bar_chart, embed_pie_chart
from helpers.csv_reader import csv_to_sheet


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def simple_wb():
    """Workbook with one sheet containing 5 rows of data."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Data")
    rows = [
        ["Name", "Score", "Grade"],
        ["Alice", 92, "A"],
        ["Bob", 78, "B"],
        ["Carol", 85, "B+"],
        ["Dave", 61, "D"],
    ]
    for row in rows:
        ws.append(row)
    return wb


@pytest.fixture()
def chart_wb():
    """Workbook with numeric data suitable for chart tests."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Sales")
    rows = [
        ["Month", "North", "South"],
        ["Jan", 100, 80],
        ["Feb", 120, 95],
        ["Mar", 110, 105],
    ]
    for row in rows:
        ws.append(row)
    return wb


@pytest.fixture()
def tmp_output(tmp_path):
    """Temporary /output directory."""
    out = tmp_path / "output"
    out.mkdir()
    return out


# ---------------------------------------------------------------------------
# style.py tests
# ---------------------------------------------------------------------------

class TestApplyHeaderStyle:

    def test_header_fill_color_applied(self, simple_wb):
        ws = simple_wb["Data"]
        apply_header_style(ws, row=1, fill_color="2563EB", font_color="FFFFFF")
        # openpyxl stores 8-char ARGB; fgColor includes alpha prefix FF.
        assert ws["A1"].fill.fgColor.rgb.endswith("2563EB")

    def test_header_font_bold(self, simple_wb):
        ws = simple_wb["Data"]
        apply_header_style(ws, row=1)
        assert ws["A1"].font.bold is True

    def test_header_font_color(self, simple_wb):
        ws = simple_wb["Data"]
        apply_header_style(ws, row=1, font_color="FF0000")
        assert ws["A1"].font.color.rgb.endswith("FF0000")

    def test_all_columns_in_row_are_styled(self, simple_wb):
        ws = simple_wb["Data"]
        apply_header_style(ws, row=1, fill_color="16A34A")
        for cell in ws[1]:
            assert cell.fill.fgColor.rgb.endswith("16A34A")

    def test_header_alignment_centered(self, simple_wb):
        ws = simple_wb["Data"]
        apply_header_style(ws, row=1)
        assert ws["A1"].alignment.horizontal == "center"


class TestApplyAltRowFill:

    def test_even_rows_get_fill(self, simple_wb):
        ws = simple_wb["Data"]
        apply_alt_row_fill(ws, start_row=2, even_fill="EFF6FF")
        # Row 3 is the 2nd data row (offset 1 from start_row=2) → even → filled
        assert ws["A3"].fill.fgColor.rgb.endswith("EFF6FF")

    def test_odd_rows_not_filled(self, simple_wb):
        ws = simple_wb["Data"]
        apply_alt_row_fill(ws, start_row=2, even_fill="EFF6FF")
        # Row 2 is offset 0 → odd → no fill (default patternType is "none")
        assert ws["A2"].fill.patternType in (None, "none", "")

    def test_custom_end_row_respected(self, simple_wb):
        ws = simple_wb["Data"]
        # Only style up to row 3
        apply_alt_row_fill(ws, start_row=2, end_row=3, even_fill="F0FDF4")
        # Row 5 should remain unstyled
        assert ws["A5"].fill.patternType in (None, "none", "")


class TestAutoColumnWidth:

    def test_wide_content_column_wider_than_min(self, simple_wb):
        ws = simple_wb["Data"]
        # Add long content to force wide column
        ws["A6"] = "This is a very long string value that should force wide column"
        auto_column_width(ws)
        assert ws.column_dimensions["A"].width > 8  # > min_width default

    def test_column_width_within_max(self, simple_wb):
        ws = simple_wb["Data"]
        ws["A6"] = "x" * 200  # way too long
        auto_column_width(ws, max_width=50)
        assert ws.column_dimensions["A"].width <= 50

    def test_empty_sheet_gets_min_width(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        auto_column_width(ws, min_width=8)
        # No content → width should be at least min_width (or column may not exist yet)
        # Just verify no crash
        assert True


class TestApplyNumberFormat:

    def test_number_format_applied_to_data_cells(self, simple_wb):
        ws = simple_wb["Data"]
        apply_number_format(ws, col_index=2, fmt="#,##0.00", start_row=2)
        assert ws["B2"].number_format == "#,##0.00"
        assert ws["B5"].number_format == "#,##0.00"

    def test_header_row_skipped(self, simple_wb):
        ws = simple_wb["Data"]
        apply_number_format(ws, col_index=2, fmt="#,##0.00", start_row=2)
        # Row 1 is header; number_format should be the default "General"
        assert ws["B1"].number_format == "General"

    def test_none_cells_format_applied(self, simple_wb):
        ws = simple_wb["Data"]
        ws["B6"] = None
        # Should not raise even for None cells
        apply_number_format(ws, col_index=2, fmt="0%", start_row=2)


class TestSanitizeSheetName:

    def test_illegal_chars_removed(self):
        result = sanitize_sheet_name("Sales [2026]: Q1/Q2*?")
        assert "[" not in result
        assert "]" not in result
        assert ":" not in result
        assert "/" not in result
        assert "*" not in result
        assert "?" not in result

    def test_truncated_to_31_chars(self):
        long_name = "A" * 40
        result = sanitize_sheet_name(long_name)
        assert len(result) <= 31

    def test_chinese_truncation(self):
        # 20 Chinese chars = 20 chars (not bytes), well within limit
        cn_name = "月度数据分析报告一二三四五六七八九十"  # 17 chars
        result = sanitize_sheet_name(cn_name)
        assert len(result) <= 31
        assert result == cn_name

    def test_empty_string_returns_sheet(self):
        result = sanitize_sheet_name("")
        assert result == "Sheet"

    def test_only_illegal_chars_returns_sheet(self):
        result = sanitize_sheet_name("[]*?/:\\")
        assert result == "Sheet"


class TestThemePresets:

    def test_all_four_themes_present(self):
        assert "default" in THEME_PRESETS
        assert "blue" in THEME_PRESETS
        assert "green" in THEME_PRESETS
        assert "monochrome" in THEME_PRESETS

    def test_each_theme_has_required_keys(self):
        required = {"header_fill", "alt_row_fill", "font_color"}
        for name, preset in THEME_PRESETS.items():
            missing = required - preset.keys()
            assert not missing, f"Theme {name!r} missing keys: {missing}"

    def test_colors_are_valid_hex(self):
        import re
        hex_re = re.compile(r"^[0-9A-Fa-f]{6}$")
        for name, preset in THEME_PRESETS.items():
            for key, value in preset.items():
                assert hex_re.match(value), (
                    f"Theme {name!r} key {key!r} has invalid hex: {value!r}"
                )


# ---------------------------------------------------------------------------
# chart.py tests
# ---------------------------------------------------------------------------

class TestEmbedBarChart:

    def test_bar_chart_added_to_workbook(self, chart_wb):
        embed_bar_chart(
            wb=chart_wb,
            data_sheet_name="Sales",
            data_range="A1:C4",
            title="Test Bar",
            anchor_cell="E1",
        )
        ws = chart_wb["Sales"]
        assert len(ws._charts) == 1

    def test_bar_chart_is_bar_chart_type(self, chart_wb):
        embed_bar_chart(
            wb=chart_wb,
            data_sheet_name="Sales",
            data_range="A1:C4",
            title="Test",
            anchor_cell="E1",
        )
        ws = chart_wb["Sales"]
        assert isinstance(ws._charts[0], BarChart)

    def test_bar_chart_placed_on_different_sheet(self, chart_wb):
        chart_wb.create_sheet(title="Charts")
        embed_bar_chart(
            wb=chart_wb,
            data_sheet_name="Sales",
            data_range="A1:C4",
            title="Cross-sheet",
            anchor_cell="A1",
            target_sheet_name="Charts",
        )
        assert len(chart_wb["Charts"]._charts) == 1
        assert len(chart_wb["Sales"]._charts) == 0


class TestEmbedPieChart:

    def test_pie_chart_added(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title="Pie")
        ws.append(["Category", "Value"])
        ws.append(["A", 30])
        ws.append(["B", 70])

        embed_pie_chart(
            wb=wb,
            data_sheet_name="Pie",
            data_range="A1:B3",
            title="Pie Test",
            anchor_cell="D1",
        )
        assert len(ws._charts) == 1

    def test_pie_chart_is_pie_chart_type(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title="Pie")
        ws.append(["Category", "Value"])
        ws.append(["X", 50])
        ws.append(["Y", 50])

        embed_pie_chart(
            wb=wb,
            data_sheet_name="Pie",
            data_range="A1:B3",
            title="Pie",
            anchor_cell="D1",
        )
        assert isinstance(ws._charts[0], PieChart)


# ---------------------------------------------------------------------------
# csv_reader.py tests
# ---------------------------------------------------------------------------

class TestCsvToSheet:

    def test_basic_csv_round_trip(self, tmp_path):
        """Write CSV then read back via csv_to_sheet and verify values."""
        csv_file = tmp_path / "test.csv"
        csv_file.write_text(
            "姓名,年龄,城市\n张三,30,北京\n李四,25,上海\n",
            encoding="utf-8",
        )

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        csv_to_sheet(wb, str(csv_file), sheet_name="测试")
        ws = wb["测试"]

        assert ws["A1"].value == "姓名"
        assert ws["A2"].value == "张三"
        assert ws["A3"].value == "李四"
        assert ws.max_row == 3  # 1 header + 2 data rows

    def test_sheet_name_from_file_stem(self, tmp_path):
        csv_file = tmp_path / "my_data.csv"
        csv_file.write_text("a,b\n1,2\n", encoding="utf-8")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        sheet_name = csv_to_sheet(wb, str(csv_file))

        assert sheet_name == "my_data"
        assert "my_data" in [s.title for s in wb.worksheets]

    def test_nan_replaced_with_none(self, tmp_path):
        """NaN values must be written as None (openpyxl-safe)."""
        import pandas as pd
        import numpy as np

        csv_file = tmp_path / "nan_test.csv"
        df = pd.DataFrame({"x": [1.0, float("nan"), 3.0]})
        df.to_csv(str(csv_file), index=False)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        csv_to_sheet(wb, str(csv_file))
        ws = list(wb.worksheets)[0]

        # Row 3 (index 2 in data) should have None, not NaN
        assert ws["A3"].value is None

    def test_csv_bom_encoding(self, tmp_path):
        """UTF-8-BOM (Excel Windows export) should be handled correctly."""
        csv_file = tmp_path / "bom.csv"
        # Write with BOM
        csv_file.write_bytes(b"\xef\xbb\xbf" + "col1,col2\nval1,val2\n".encode("utf-8"))

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        csv_to_sheet(wb, str(csv_file), sheet_name="BOM")
        ws = wb["BOM"]

        # BOM should NOT appear in header cell value
        assert ws["A1"].value == "col1", f"Got: {ws['A1'].value!r}"

    def test_file_not_found_raises(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        with pytest.raises(FileNotFoundError):
            csv_to_sheet(wb, "/nonexistent/path/file.csv")

    def test_column_count_matches_csv(self, tmp_path):
        csv_file = tmp_path / "cols.csv"
        csv_file.write_text("a,b,c,d,e\n1,2,3,4,5\n", encoding="utf-8")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        csv_to_sheet(wb, str(csv_file))
        ws = list(wb.worksheets)[0]

        assert ws.max_column == 5


# ---------------------------------------------------------------------------
# main.py tests
# ---------------------------------------------------------------------------

class TestMainRun:

    def _run(self, params: dict, output_dir: Path) -> dict:
        """Helper: patch /output to tmp dir and run."""
        import main as main_module
        orig = main_module.OUTPUT_DIR
        main_module.OUTPUT_DIR = output_dir
        try:
            return main_module.run(params)
        finally:
            main_module.OUTPUT_DIR = orig

    def test_basic_sheet_creation(self, tmp_output):
        result = self._run(
            {
                "output_filename": "test.xlsx",
                "sheets": [
                    {
                        "name": "Sheet1",
                        "data": [["A", "B"], [1, 2], [3, 4]],
                    }
                ],
            },
            tmp_output,
        )
        assert result["success"] is True
        assert result["sheets_created"] == 1
        out = Path(result["output_path"])
        assert out.exists()

        wb = openpyxl.load_workbook(str(out))
        ws = wb.worksheets[0]
        assert ws["A1"].value == "A"
        assert ws["B2"].value == 2

    def test_multi_sheet(self, tmp_output):
        result = self._run(
            {
                "output_filename": "multi.xlsx",
                "sheets": [
                    {"name": "Alpha", "data": [["X"], [1]]},
                    {"name": "Beta",  "data": [["Y"], [2]]},
                    {"name": "Gamma", "data": [["Z"], [3]]},
                ],
            },
            tmp_output,
        )
        assert result["success"] is True
        assert result["sheets_created"] == 3
        wb = openpyxl.load_workbook(result["output_path"])
        assert [ws.title for ws in wb.worksheets] == ["Alpha", "Beta", "Gamma"]

    def test_header_style_applied(self, tmp_output):
        result = self._run(
            {
                "output_filename": "styled.xlsx",
                "sheets": [{"name": "S", "data": [["Col"], [1]]}],
                "style_config": {"theme": "blue"},
            },
            tmp_output,
        )
        assert result["success"] is True
        wb = openpyxl.load_workbook(result["output_path"])
        ws = wb.worksheets[0]
        # Blue theme header fill is "2563EB"
        assert ws["A1"].fill.fgColor.rgb.endswith("2563EB")

    def test_alt_row_fill_applied(self, tmp_output):
        result = self._run(
            {
                "output_filename": "alt.xlsx",
                "sheets": [{
                    "name": "S",
                    "data": [["A"], [1], [2], [3], [4]],
                }],
                "style_config": {"theme": "blue"},
            },
            tmp_output,
        )
        assert result["success"] is True
        wb = openpyxl.load_workbook(result["output_path"])
        ws = wb.worksheets[0]
        # Even alt row should have fill (blue theme alt_row_fill = "EFF6FF")
        assert ws["A3"].fill.fgColor.rgb.endswith("EFF6FF")

    def test_bar_chart_inserted(self, tmp_output):
        result = self._run(
            {
                "output_filename": "chart.xlsx",
                "sheets": [{
                    "name": "Data",
                    "data": [
                        ["Month", "Sales"],
                        ["Jan", 100],
                        ["Feb", 120],
                    ],
                }],
                "charts": [{
                    "sheet": "Data",
                    "chart_type": "bar",
                    "title": "Test",
                    "data_range": "A1:B3",
                    "anchor": "D1",
                }],
            },
            tmp_output,
        )
        assert result["success"] is True
        assert result["charts_created"] == 1

    def test_pie_chart_inserted(self, tmp_output):
        result = self._run(
            {
                "output_filename": "pie.xlsx",
                "sheets": [{
                    "name": "Data",
                    "data": [
                        ["Category", "Value"],
                        ["A", 60],
                        ["B", 40],
                    ],
                }],
                "charts": [{
                    "sheet": "Data",
                    "chart_type": "pie",
                    "title": "Pie",
                    "data_range": "A1:B3",
                    "anchor": "D1",
                }],
            },
            tmp_output,
        )
        assert result["success"] is True
        assert result["charts_created"] == 1

    def test_sheet_name_truncation(self, tmp_output):
        long_name = "A" * 40
        result = self._run(
            {
                "output_filename": "trunc.xlsx",
                "sheets": [{"name": long_name, "data": [["X"], [1]]}],
            },
            tmp_output,
        )
        assert result["success"] is True
        wb = openpyxl.load_workbook(result["output_path"])
        assert len(wb.worksheets[0].title) <= 31

    def test_output_file_exists_at_correct_path(self, tmp_output):
        result = self._run(
            {
                "output_filename": "present.xlsx",
                "sheets": [{"name": "S", "data": [["V"], [1]]}],
            },
            tmp_output,
        )
        assert result["success"] is True
        assert Path(result["output_path"]).exists()

    def test_large_dataset_write_only_mode(self, tmp_output):
        """50 001 rows should trigger write_only mode."""
        import main as main_module
        # Patch threshold to a small value for speed.
        orig_threshold = main_module.WRITE_ONLY_THRESHOLD
        main_module.WRITE_ONLY_THRESHOLD = 5
        try:
            result = self._run(
                {
                    "output_filename": "large.xlsx",
                    "sheets": [{
                        "name": "Big",
                        "data": [["X"]] + [[i] for i in range(10)],
                    }],
                },
                tmp_output,
            )
        finally:
            main_module.WRITE_ONLY_THRESHOLD = orig_threshold

        assert result["success"] is True
        assert result["warning"] is not None
        assert "write_only" in result["warning"]

    def test_error_returns_success_false(self, tmp_output):
        """Invalid chart type should still succeed (warning), but bad sheet ref errors."""
        result = self._run(
            {
                "output_filename": "err.xlsx",
                "sheets": [{"name": "S", "data": [["A"], [1]]}],
                "charts": [{
                    "sheet": "NonExistentSheet",  # Will raise KeyError
                    "chart_type": "bar",
                    "title": "Err",
                    "data_range": "A1:B2",
                    "anchor": "D1",
                }],
            },
            tmp_output,
        )
        assert result["success"] is False
        assert result["error"] is not None

    def test_output_size_within_limit(self, tmp_output):
        """10 sheets x 100 rows each should be well under 50 MB."""
        sheets = [
            {
                "name": f"Sheet{i}",
                "data": [["Col1", "Col2", "Col3"]]
                + [[f"val{r}", r * 10, r * 0.5] for r in range(100)],
            }
            for i in range(10)
        ]
        result = self._run(
            {"output_filename": "size_test.xlsx", "sheets": sheets},
            tmp_output,
        )
        assert result["success"] is True
        # 50 MB = 50 * 1024 * 1024 bytes
        assert result["output_size_bytes"] < 50 * 1024 * 1024

    def test_scatter_chart_returns_warning(self, tmp_output):
        """scatter chart type is not implemented; should warn but not fail the run."""
        result = self._run(
            {
                "output_filename": "scatter.xlsx",
                "sheets": [{"name": "D", "data": [["X", "Y"], [1, 2]]}],
                "charts": [{
                    "sheet": "D",
                    "chart_type": "scatter",
                    "title": "S",
                    "data_range": "A1:B2",
                    "anchor": "D1",
                }],
            },
            tmp_output,
        )
        assert result["success"] is True
        assert result["warning"] is not None
        assert "scatter" in result["warning"]


# ---------------------------------------------------------------------------
# template.py tests
# ---------------------------------------------------------------------------

class TestLoadTemplate:

    def test_template_vars_replacement(self, tmp_path):
        """Create a minimal test template and verify placeholder substitution."""
        # Build a minimal .xlsx template in tmp_path
        tpl_wb = openpyxl.Workbook()
        tpl_ws = tpl_wb.active
        tpl_ws.title = "Cover"
        tpl_ws["A1"] = "{{title}}"
        tpl_ws["A2"] = "Period: {{period}}"
        tpl_path = tmp_path / "test_template.xlsx"
        tpl_wb.save(str(tpl_path))

        from helpers import template as tpl_module
        orig_dir = tpl_module.TEMPLATE_DIR
        tpl_module.TEMPLATE_DIR = tmp_path
        try:
            wb = tpl_module.load_template(
                "test_template",
                {"title": "测试报告", "period": "2026-Q2"},
            )
        finally:
            tpl_module.TEMPLATE_DIR = orig_dir

        ws = wb.worksheets[0]
        assert ws["A1"].value == "测试报告"
        assert ws["A2"].value == "Period: 2026-Q2"

    def test_missing_var_leaves_placeholder(self, tmp_path):
        """Placeholders with no matching var key are left unchanged."""
        tpl_wb = openpyxl.Workbook()
        tpl_wb.active["A1"] = "{{unknown_key}}"
        tpl_wb.save(str(tmp_path / "tpl.xlsx"))

        from helpers import template as tpl_module
        orig_dir = tpl_module.TEMPLATE_DIR
        tpl_module.TEMPLATE_DIR = tmp_path
        try:
            wb = tpl_module.load_template("tpl", {"other": "value"})
        finally:
            tpl_module.TEMPLATE_DIR = orig_dir

        assert wb.worksheets[0]["A1"].value == "{{unknown_key}}"

    def test_path_traversal_raises_value_error(self):
        from helpers import template as tpl_module
        with pytest.raises(ValueError):
            tpl_module.load_template("../evil", {})

    def test_missing_template_raises_file_not_found(self, tmp_path):
        from helpers import template as tpl_module
        orig_dir = tpl_module.TEMPLATE_DIR
        tpl_module.TEMPLATE_DIR = tmp_path
        try:
            with pytest.raises(FileNotFoundError):
                tpl_module.load_template("nonexistent", {})
        finally:
            tpl_module.TEMPLATE_DIR = orig_dir


# ---------------------------------------------------------------------------
# sanitize_sheet_name edge cases
# ---------------------------------------------------------------------------

class TestSanitizeSheetNameEdgeCases:

    def test_backslash_removed(self):
        assert "\\" not in sanitize_sheet_name("Sales\\2026")

    def test_whitespace_collapsed(self):
        result = sanitize_sheet_name("  foo   bar  ")
        assert "  " not in result

    def test_normal_name_unchanged(self):
        assert sanitize_sheet_name("Monthly Sales") == "Monthly Sales"

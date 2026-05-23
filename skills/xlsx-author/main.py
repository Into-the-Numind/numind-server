"""
main.py — xlsx-author skill entry point.

Called by the invoke_skill framework with a JSON-serialisable params dict.
Reads params, builds an openpyxl Workbook (optionally from a template),
writes sheets and charts, saves to /output/, and returns a structured result.

Input params schema (all fields optional unless noted):
    output_filename  (required) str  -- e.g. "报告.xlsx"
    sheets           list[dict]      -- sheet definitions
    csv_input_path   str             -- path under /workspace/input/ to append
    charts           list[dict]      -- chart definitions
    style_config     dict            -- colour theme / font overrides
    template         str             -- "summary" | "daily-report" | "comparison"
    template_vars    dict            -- placeholder → replacement for template

Output dict:
    success          bool
    output_path      str
    output_size_bytes int
    sheets_created   int
    charts_created   int
    warning          str | None
    error            str | None

Usage (from sandbox):
    import json, sys
    params = json.load(sys.stdin)  # OR passed by invoke_skill framework
    from main import run
    result = run(params)
    print(json.dumps(result))
"""

from __future__ import annotations

import os
import re
import sys
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.workbook.workbook import Workbook

# Adjust sys.path so helpers are importable when main.py is run directly.
_SKILL_DIR = Path(__file__).parent
if str(_SKILL_DIR) not in sys.path:
    sys.path.insert(0, str(_SKILL_DIR))

from helpers import (
    apply_header_style,
    apply_alt_row_fill,
    auto_column_width,
    apply_number_format,
    THEME_PRESETS,
)
from helpers.chart import embed_bar_chart, embed_line_chart, embed_pie_chart
from helpers.csv_reader import csv_to_sheet
from helpers.template import load_template
from helpers.style import sanitize_sheet_name

OUTPUT_DIR = Path("/output")
INPUT_DIR = Path("/workspace/input")

# Rows threshold above which write_only mode is engaged (charts disabled).
WRITE_ONLY_THRESHOLD = 50_000

# Sanitise output filename: keep alphanumerics, CJK, hyphens, underscores, dots.
_SAFE_FILENAME_RE = re.compile(r"[^\w一-鿿぀-ヿ\-_.]", re.UNICODE)


def _sanitize_filename(name: str) -> str:
    """Remove unsafe characters and ensure a .xlsx extension."""
    # Force basename (no path traversal).
    name = Path(name).name
    # Replace unsafe characters with underscore.
    name = _SAFE_FILENAME_RE.sub("_", name)
    if not name:
        name = "output"
    if not name.lower().endswith(".xlsx"):
        name = name + ".xlsx"
    return name


def _build_workbook_normal(params: dict[str, Any]) -> tuple[Workbook, int, int, Optional[str]]:
    """Build a Workbook in normal (non-write_only) mode.

    Returns (wb, sheets_created, charts_created, warning_or_None).
    """
    style = params.get("style_config", {})
    theme_name = style.get("theme", "default")
    theme = THEME_PRESETS.get(theme_name, THEME_PRESETS["default"])

    header_fill = style.get("header_fill", theme["header_fill"]).lstrip("#")
    font_color = theme.get("font_color", "FFFFFF")
    alt_row_fill = style.get("alt_row_fill", theme["alt_row_fill"]).lstrip("#")
    font_name = style.get("font_name", "微软雅黑")
    font_size = int(style.get("font_size", 10))
    number_fmt = style.get("number_format", "#,##0.00")
    date_fmt = style.get("date_format", "YYYY-MM-DD")

    # Start from template if requested, otherwise empty workbook.
    template_name = params.get("template")
    if template_name:
        template_vars = params.get("template_vars", {})
        wb = load_template(template_name, template_vars)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default "Sheet"

    sheets_created = 0

    # Process sheet definitions.
    for sheet_def in params.get("sheets", []):
        raw_name = sheet_def.get("name", f"Sheet{sheets_created + 1}")
        ws = wb.create_sheet(title=sanitize_sheet_name(raw_name))

        data = sheet_def.get("data", [])
        for row_idx, row in enumerate(data, start=1):
            for col_idx, val in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        has_headers = sheet_def.get("headers", True)
        if has_headers and data:
            apply_header_style(
                ws, row=1,
                fill_color=header_fill,
                font_color=font_color,
                font_name=font_name,
                font_size=font_size,
            )
            apply_alt_row_fill(ws, start_row=2, even_fill=alt_row_fill)

        if sheet_def.get("freeze_top_row", True) and data:
            ws.freeze_panes = "A2"

        # Apply column widths if specified, otherwise auto-size.
        col_widths = sheet_def.get("col_widths")
        if col_widths:
            from openpyxl.utils import get_column_letter
            for i, width in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = int(width)
        else:
            auto_column_width(ws)

        sheets_created += 1

    # Append CSV input if provided.
    csv_input = params.get("csv_input_path")
    if csv_input:
        full_csv = INPUT_DIR / Path(csv_input).name
        sheet_name = csv_to_sheet(wb, str(full_csv))
        # Apply basic styling to the CSV sheet.
        ws_csv = wb[sheet_name]
        if ws_csv.max_row > 1:
            apply_header_style(
                ws_csv, row=1,
                fill_color=header_fill,
                font_color=font_color,
                font_name=font_name,
            )
            apply_alt_row_fill(ws_csv, start_row=2, even_fill=alt_row_fill)
            auto_column_width(ws_csv)
        sheets_created += 1

    # Process charts.
    warning = None
    charts_created = 0
    chart_dispatch = {
        "bar": embed_bar_chart,
        "line": embed_line_chart,
        "pie": embed_pie_chart,
    }

    for chart_def in params.get("charts", []):
        chart_type = chart_def.get("chart_type", "bar")
        if chart_type == "scatter":
            warning = (
                "scatter chart type is not yet implemented (V1.5); "
                "skipped. Use bar/line/pie or chart_from_matplotlib."
            )
            continue

        embed_fn = chart_dispatch.get(chart_type)
        if embed_fn is None:
            warning = f"Unknown chart_type: {chart_type!r}; skipped."
            continue

        kwargs: dict[str, Any] = dict(
            wb=wb,
            data_sheet_name=chart_def["sheet"],
            data_range=chart_def["data_range"],
            title=chart_def.get("title", ""),
            anchor_cell=chart_def.get("anchor", "F2"),
            target_sheet_name=chart_def.get("place_on_sheet"),
        )
        # embed_pie_chart does not accept series_labels.
        if chart_type != "pie" and chart_def.get("series_labels"):
            kwargs["series_labels"] = chart_def["series_labels"]

        embed_fn(**kwargs)
        charts_created += 1

    return wb, sheets_created, charts_created, warning


def _build_workbook_write_only(params: dict[str, Any]) -> tuple[Workbook, int, Optional[str]]:
    """Build a Workbook in write_only mode for large datasets (> 50 000 rows).

    Charts are not supported in write_only mode and will be silently skipped.
    Returns (wb, sheets_created, warning).
    """
    wb = openpyxl.Workbook(write_only=True)
    sheets_created = 0

    for sheet_def in params.get("sheets", []):
        raw_name = sheet_def.get("name", f"Sheet{sheets_created + 1}")
        ws = wb.create_sheet(title=sanitize_sheet_name(raw_name))

        data = sheet_def.get("data", [])
        for row in data:
            ws.append(row)

        sheets_created += 1

    warning = "write_only mode engaged (>50 000 rows): charts and styles are skipped."
    return wb, sheets_created, warning


def run(params: dict[str, Any]) -> dict[str, Any]:
    """Main entry point called by the invoke_skill framework.

    Args:
        params: Deserialised invoke params dict.  Must contain
                ``output_filename``.

    Returns:
        Result dict with keys: success, output_path, output_size_bytes,
        sheets_created, charts_created, warning, error.
    """
    try:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        output_filename = _sanitize_filename(params.get("output_filename", "output.xlsx"))
        output_path = OUTPUT_DIR / output_filename

        # Determine total row count to decide mode.
        total_rows = sum(
            len(sd.get("data", [])) for sd in params.get("sheets", [])
        )

        if total_rows > WRITE_ONLY_THRESHOLD:
            wb, sheets_created, warning = _build_workbook_write_only(params)
            charts_created = 0
        else:
            wb, sheets_created, charts_created, warning = _build_workbook_normal(params)

        wb.save(str(output_path))
        size_bytes = output_path.stat().st_size

        return {
            "success": True,
            "output_path": str(output_path),
            "output_size_bytes": size_bytes,
            "sheets_created": sheets_created,
            "charts_created": charts_created,
            "warning": warning,
            "error": None,
        }

    except Exception as exc:
        return {
            "success": False,
            "output_path": None,
            "output_size_bytes": 0,
            "sheets_created": 0,
            "charts_created": 0,
            "warning": None,
            "error": str(exc),
        }


if __name__ == "__main__":
    import json

    raw = sys.stdin.read().strip()
    if raw:
        p = json.loads(raw)
    else:
        # Minimal smoke-test when called without input.
        p = {
            "output_filename": "smoke_test.xlsx",
            "sheets": [
                {
                    "name": "测试Sheet",
                    "data": [["姓名", "分数"], ["张三", 90], ["李四", 85]],
                }
            ],
        }

    result = run(p)
    print(json.dumps(result, ensure_ascii=False, indent=2))
